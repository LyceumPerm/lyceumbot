from openpyxl import load_workbook

from constants import CURRENT_TABLE, CURRENT_FILE, available_days, alt_teachers, alt_profiles
from database import ScheduleTable


class TableParser:
    def __init__(self, file_path: str):
        self.wb = load_workbook(file_path)
        self.sheet = self.wb[CURRENT_TABLE]
        self.schedule_db = ScheduleTable()

    def parse(self):
        merged_cells = self.sheet.merged_cells

        for clas in range(3, 76, 3):
            if clas == 39:
                continue
            if clas > 39:
                clas -= 1

            for day in range(4, 25, 5):
                date = self.sheet.cell(day, 1).value.strftime('%d.%m')

                # баг в таблице
                if date == '02.02':
                    date = '02.03'

                for row in range(day, day + 5):

                    merged = False  # true - общая пара; false - по группам
                    for i in str(merged_cells).split():
                        if str(i).startswith(self.sheet.cell(row, clas).coordinate):
                            merged = True

                    table_class = self.sheet.cell(2, clas).value
                    class_number = table_class[:2]
                    class_profile = table_class[2:].lower()
                    if class_profile in alt_profiles:
                        class_profile = alt_profiles[class_profile]

                    number = row - day + 1
                    classroom = str(self.sheet.cell(row, clas + 2).value)
                    if classroom.endswith('.0'):
                        classroom = classroom[:-2]

                    if merged:
                        text = self.sheet.cell(row, clas).value
                        try:
                            subject = text[:text.index('(')].strip()
                            teacher = text[text.index('(') + 1:text.index(')')]
                        except:
                            subject = text
                            teacher = ''

                        teacher = self.format_name(teacher)

                        if self.sheet.cell(row, clas).font.strikethrough:
                            try:
                                subject = '<s>' + subject + '</s>'
                                teacher = '<s>' + teacher + '</s>'
                            except:
                                pass

                        self.schedule_db.save(date, number, subject, teacher, class_number, class_profile, 1, classroom)
                        self.schedule_db.save(date, number, subject, teacher, class_number, class_profile, 2, classroom)

                    else:
                        text1 = self.sheet.cell(row, clas).value
                        try:
                            subject1 = text1[:text1.index('(')].strip()
                            teacher1 = text1[text1.index('(') + 1:text1.index(')')]
                        except:
                            subject1 = text1
                            teacher1 = ''

                        text2 = self.sheet.cell(row, clas + 1).value
                        try:
                            subject2 = text2[:text2.index('(')].strip()
                            teacher2 = text2[text2.index('(') + 1:text2.index(')')]
                        except:
                            subject2 = text2
                            teacher2 = ''

                        teacher1 = self.format_name(teacher1)
                        teacher2 = self.format_name(teacher2)

                        if self.sheet.cell(row, clas).font.strikethrough:
                            try:
                                subject1 = '<s>' + subject1 + '</s>'
                                teacher1 = '<s>' + teacher1 + '</s>'
                            except:
                                pass
                        if self.sheet.cell(row, clas + 1).font.strikethrough:
                            try:
                                subject2 = '<s>' + subject2 + '</s>'
                                teacher2 = '<s>' + teacher2 + '</s>'
                            except:
                                pass

                        classrooms = [classroom, classroom]
                        if '/' in classroom and classroom.lower() not in ['сп/з', 'с/з', 'а/з', 'акт/з', 'сп/зал',
                                                                          'с/зал', 'а/зал', 'акт/зал']:
                            classrooms = classroom.split('/')

                        self.schedule_db.save(date, number, subject1, teacher1, class_number, class_profile, 1, classrooms[0])
                        self.schedule_db.save(date, number, subject2, teacher2, class_number, class_profile, 2, classrooms[1])

    def format_name(self, name: str) -> str:
        if len(name) < 3:
            return name
        if name.lower() in alt_teachers:
            name = alt_teachers[name.lower()]
        name = name.replace('.', '').replace(' ', '')
        return f'{name[:-2]} {name[-2]}.{name[-1]}.'

    def clear(self, days_to_delete):
        self.schedule_db.clear(days_to_delete)

    def __del__(self):
        self.wb.close()
        self.schedule_db.con.close()


if __name__ == '__main__':
    parser = TableParser(f'resources/schedule/{CURRENT_FILE}')
    parser.clear(available_days[-5:])
    parser.parse()
    parser.__del__()
