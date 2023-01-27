from openpyxl import load_workbook
# from constants import table_indexes
from exceptions import ClasException, GroupException

table_indexes = {'10б': 'B2', '10геохим': 'E2', '10исс1': 'H2', '10исс2': 'K2', '10линг1': 'N2', '10линг2': 'B8',
                 '10мат': 'E8', '10матэк': 'H8', '10мед': 'K8', '10фил': 'N8', '10фм': 'B14', '10эк': 'E14',
                 '11б': 'H14', '11бх': 'K14', '11гео': 'N14', '11исс': 'B20', '11линг1': 'E20', '11линг2': 'H20',
                 '11м1': 'K20', '11м2': 'N20', '11мед': 'B26', '11фил': 'E26', '11фм': 'H26', '11эк': 'K26'}


# old parser
class Table:
    def __init__(self, date: str):
        self.wb = load_workbook(f'resources/schedule/{date}.xlsx')
        self.sheet = self.wb[self.wb.sheetnames[0]]

    def get_raw(self, date: str, clas: str):
        if clas not in table_indexes:
            raise ClasException

        # клетка с названием класса
        title_cell = self.sheet[table_indexes[clas]]
        title_row = title_cell.row
        title_column = title_cell.column

        merged_cells = self.sheet.merged_cells

        # заполнение пар по группам
        group1 = []
        group2 = []
        for row in range(title_row + 1, title_row + 6):
            merged = False
            for i in str(merged_cells).split():
                if str(i).startswith(self.sheet.cell(row, title_column).coordinate):
                    merged = True
            if merged:
                group1.append(self.sheet.cell(row, title_column).value)
                group2.append(self.sheet.cell(row, title_column).value)
            else:
                group1.append(self.sheet.cell(row, title_column).value)
                group2.append(self.sheet.cell(row, title_column + 1).value)

        # заполнение кабинетов
        classrooms1 = []
        classrooms2 = []
        for row in range(title_row + 1, title_row + 6):
            cell_text = str(self.sheet.cell(row, title_column + 2).value)
            if '/' in cell_text and cell_text.lower() not in ['сп/з', 'с/з', 'а/з', 'акт/з', 'сп/зал', 'с/зал', 'а/зал',
                                                              'акт/зал']:
                classrooms1.append(cell_text.split('/')[0])
                classrooms2.append(cell_text.split('/')[1])
            else:
                classrooms1.append(cell_text)
                classrooms2.append(cell_text)

        # удаляет пятую пару, если она None
        if group1[4] is None and group2[4] is None:
            group1.pop(4)
            group2.pop(4)
            classrooms1.pop(4)
            classrooms2.pop(4)

        return group1, group2, classrooms1, classrooms2

    def get_schedule(self, date: str, clas: str, group: int):
        if group not in [1, 2]:
            raise GroupException
        return self.get_raw(date, clas)[group - 1]

    def get_classrooms(self, date: str, clas: str, group: int):
        if group not in [1, 2]:
            raise GroupException
        return self.get_raw(date, clas)[1 + group]

    def set_workbook(self, wb_path: str):
        self.wb = load_workbook(wb_path)
